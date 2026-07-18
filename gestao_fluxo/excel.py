"""Geração das planilhas de download — relatório executivo em .xlsx.

Recebe o DataFrame *cru* (número como número, data como data) e devolve os bytes
da pasta de trabalho. A formatação vive aqui, não nos valores: quem chama não
precisa pré-formatar nada em texto, e o Excel continua somando e ordenando.

Paleta clara e sóbria (o tema escuro do app não se traduz bem em planilha
impressa): faixa de título em grafite, cabeçalho grafite com fio de acento
turquesa, linhas alternadas em cinza muito claro.
"""
from __future__ import annotations

import re
from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

from .exceptions import RelatorioError

# Cores da planilha (ARGB sem o "#", como o openpyxl espera).
_GRAFITE = "FF0F172A"      # faixa de título e cabeçalho
_ACENTO = "FF2DD4BF"       # turquesa da marca, só no fio do cabeçalho
_TINTA = "FF0F172A"        # texto dos dados
_MUTED = "FF64748B"        # subtítulo e contexto
_ZEBRA = "FFF6F8FA"        # faixa alternada das linhas
_LINHA = "FFE2E8F0"        # borda fina entre linhas
_TOTAL = "FFEDF7F5"        # fundo da linha de totais

_FONTE = "Segoe UI"

# Formatos numéricos: o Excel troca ponto/vírgula conforme o idioma da máquina,
# então "#,##0" já sai como 1.234 num Windows em pt-BR.
_FMT_INT = "#,##0"
_FMT_DEC = "#,##0.00"
_FMT_DATA = "dd/mm/yyyy"

# Altura da faixa de título, em linhas, antes do cabeçalho da tabela.
_LINHAS_TITULO = 3

# Caracteres de controle que o Excel recusa — o openpyxl levanta
# IllegalCharacterError e derrubaria a tela inteira no meio de um download.
# Eles chegam do export sujo da origem: `limpar_texto` do ETL colapsa espaços,
# mas \x01 e afins não são espaço e passam batido até aqui.
_ILEGAIS = re.compile(r"[\000-\010\013\014\016-\037]")


def _limpo(valor):
    """Tira do texto o que o Excel não aceita; qualquer outro tipo passa direto."""
    return _ILEGAIS.sub("", valor) if isinstance(valor, str) else valor


def _formato(serie: pd.Series) -> str | None:
    """Formato de exibição da coluna, deduzido do dtype. None = texto."""
    if pd.api.types.is_datetime64_any_dtype(serie):
        return _FMT_DATA
    if pd.api.types.is_bool_dtype(serie):
        return None
    if pd.api.types.is_integer_dtype(serie):
        return _FMT_INT
    if pd.api.types.is_float_dtype(serie):
        # Coluna float só com valores inteiros (soma de peças, p. ex.) não
        # precisa de casas decimais penduradas.
        limpa = serie.dropna()
        if not limpa.empty and (limpa % 1 == 0).all():
            return _FMT_INT
        return _FMT_DEC
    return None


def _valores(serie: pd.Series) -> list:
    """Coluna inteira convertida de uma vez para tipos nativos do Python.

    Vale a conversão em bloco: percorrer a base com `iterrows()` custa ordens de
    grandeza mais e este caminho roda na base completa, não na página exibida.
    Ausente (NaN/NaT) vira None — célula vazia lê melhor que um traço no Excel.
    """
    nulos = serie.isna().to_numpy()
    if pd.api.types.is_datetime64_any_dtype(serie):
        bruto = serie.dt.to_pydatetime()
    else:
        bruto = serie.to_numpy(dtype=object)
    return [None if vazio else _limpo(valor) for valor, vazio in zip(bruto, nulos)]


def _largura(serie: pd.Series, rotulo: str, formato: str | None) -> float:
    """Largura da coluna pelo maior conteúdo, com piso e teto para não distorcer."""
    if formato == _FMT_DATA:
        maior = 10
    else:
        # `head` antes do `astype`: converter a coluna inteira para texto só para
        # ler as 400 primeiras desperdiça uma conversão por linha da base.
        amostra = serie.dropna().head(400).astype(str)
        maior = int(amostra.map(len).max()) if not amostra.empty else 0
        if formato in (_FMT_INT, _FMT_DEC):
            maior += 2                # espaço dos separadores de milhar
    return min(max(len(rotulo) + 4, maior + 4), 46)


def _nome_aba(titulo: str) -> str:
    """Nome de aba válido: sem os caracteres proibidos e no limite de 31 chars."""
    limpo = re.sub(r"[\[\]:*?/\\]", " ", titulo).strip()
    return (limpo or "Relatório")[:31]


def gerar_xlsx(df: pd.DataFrame, colunas: dict, *, titulo: str,
               subtitulo: str = "", somar: tuple = ()) -> bytes:
    """Monta o relatório executivo e devolve os bytes do .xlsx.

    colunas: {chave_do_df: 'Rótulo exibido'}, na ordem desejada.
    somar: chaves que ganham uma linha de TOTAL ao pé da tabela.

    Qualquer falha de montagem vira `RelatorioError`. O download é um extra da
    tela: perdê-lo não pode derrubar a tabela que responde a pergunta do operador.
    """
    try:
        return _montar(df, colunas, titulo, subtitulo, somar)
    except Exception as exc:  # noqa: BLE001
        raise RelatorioError(f"Falha ao gerar a planilha '{titulo}': {exc}") from exc


def _montar(df: pd.DataFrame, colunas: dict, titulo: str, subtitulo: str,
            somar: tuple) -> bytes:
    """Monta a pasta de trabalho de fato — sem tratamento, quem chama traduz."""
    dados = df[list(colunas)].copy()
    formatos = {chave: _formato(dados[chave]) for chave in colunas}
    n_col = len(colunas)

    wb = Workbook()
    ws = wb.active
    ws.title = _nome_aba(titulo)
    ws.sheet_view.showGridLines = False   # a grade compete com as bordas do relatório
    # Altura no padrão da planilha em vez de linha a linha: numa base de dezenas
    # de milhares de registros, um objeto de dimensão por linha domina o custo.
    ws.sheet_format.defaultRowHeight = 19
    ws.sheet_format.customHeight = True

    _faixa_titulo(ws, titulo, subtitulo, len(dados), n_col)
    lin_cab = _LINHAS_TITULO + 1
    _cabecalho(ws, colunas, formatos, lin_cab)
    ultima = _corpo(ws, dados, colunas, formatos, lin_cab + 1)
    if somar and not dados.empty:
        ultima = _linha_total(ws, dados, colunas, formatos, somar, ultima + 1)

    _acabamento(ws, dados, colunas, formatos, lin_cab, ultima, n_col)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _faixa_titulo(ws, titulo: str, subtitulo: str, linhas: int, n_col: int) -> None:
    """Bloco de abertura: título em grafite e uma linha de contexto abaixo."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_col)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_col)

    cel = ws.cell(row=1, column=1, value=titulo)
    cel.font = Font(name=_FONTE, size=16, bold=True, color=_GRAFITE)
    cel.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30

    contexto = f"{linhas:,} registro(s)".replace(",", ".")
    if subtitulo:
        contexto = f"{subtitulo}  ·  {contexto}"
    contexto += f"  ·  Gerado em {datetime.now():%d/%m/%Y %H:%M}"
    cel = ws.cell(row=2, column=1, value=contexto)
    cel.font = Font(name=_FONTE, size=9, color=_MUTED)
    cel.alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 16
    ws.row_dimensions[_LINHAS_TITULO].height = 8   # respiro antes da tabela


def _cabecalho(ws, colunas: dict, formatos: dict, linha: int) -> None:
    """Cabeçalho grafite com o fio turquesa embaixo."""
    fundo = PatternFill("solid", fgColor=_GRAFITE)
    fonte = Font(name=_FONTE, size=10, bold=True, color="FFFFFFFF")
    fio = Border(bottom=Side(style="medium", color=_ACENTO))
    for i, (chave, rotulo) in enumerate(colunas.items(), start=1):
        cel = ws.cell(row=linha, column=i, value=rotulo)
        cel.fill, cel.font, cel.border = fundo, fonte, fio
        cel.alignment = Alignment(
            horizontal="right" if formatos[chave] else "left",
            vertical="center", wrap_text=True,
        )
    ws.row_dimensions[linha].height = 26


def _corpo(ws, dados: pd.DataFrame, colunas: dict, formatos: dict,
           primeira: int) -> int:
    """Escreve as linhas com zebra e devolve o número da última linha.

    Os objetos de estilo são criados uma vez e reaproveitados em todas as
    células: o openpyxl deduplica estilos, então compartilhar é mais rápido e
    ainda gera um arquivo menor.
    """
    zebra = PatternFill("solid", fgColor=_ZEBRA)
    fonte = Font(name=_FONTE, size=10, color=_TINTA)
    borda = Border(bottom=Side(style="thin", color=_LINHA))
    alinha = [
        Alignment(horizontal="right" if formatos[chave] else "left",
                  vertical="center")
        for chave in colunas
    ]
    fmts = [formatos[chave] for chave in colunas]

    linha = primeira - 1
    colunas_vals = [_valores(dados[chave]) for chave in colunas]
    for n, registro in enumerate(zip(*colunas_vals)):
        linha = primeira + n
        listrada = n % 2
        for i, valor in enumerate(registro):
            cel = ws.cell(row=linha, column=i + 1, value=valor)
            cel.font, cel.border, cel.alignment = fonte, borda, alinha[i]
            if fmts[i]:
                cel.number_format = fmts[i]
            if listrada:
                cel.fill = zebra
    return linha


def _linha_total(ws, dados: pd.DataFrame, colunas: dict, formatos: dict,
                 somar: tuple, linha: int) -> int:
    """Fecha a tabela com os totais das colunas pedidas."""
    fundo = PatternFill("solid", fgColor=_TOTAL)
    fonte = Font(name=_FONTE, size=10, bold=True, color=_TINTA)
    borda = Border(top=Side(style="medium", color=_ACENTO))
    for i, chave in enumerate(colunas, start=1):
        valor = None
        if chave in somar:
            valor = float(dados[chave].sum())
            if formatos[chave] == _FMT_INT:
                valor = int(valor)
        cel = ws.cell(row=linha, column=i, value="TOTAL" if i == 1 else valor)
        cel.fill, cel.font, cel.border = fundo, fonte, borda
        cel.alignment = Alignment(
            horizontal="left" if i == 1 else "right", vertical="center")
        if chave in somar and formatos[chave]:
            cel.number_format = formatos[chave]
    ws.row_dimensions[linha].height = 22
    return linha


def _acabamento(ws, dados: pd.DataFrame, colunas: dict, formatos: dict,
                lin_cab: int, ultima: int, n_col: int) -> None:
    """Larguras, painel congelado, filtro e preparo de impressão."""
    for i, (chave, rotulo) in enumerate(colunas.items(), start=1):
        ws.column_dimensions[get_column_letter(i)].width = _largura(
            dados[chave], rotulo, formatos[chave])

    ws.freeze_panes = ws.cell(row=lin_cab + 1, column=1)
    if not dados.empty:
        ws.auto_filter.ref = f"A{lin_cab}:{get_column_letter(n_col)}{ultima}"

    ws.print_title_rows = f"{lin_cab}:{lin_cab}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)
